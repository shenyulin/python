import os
import xlwt

# 学习时间：20200609
# ------1、快速整理一个盘或文件夹下的文件名称列表 -------#
file_dir = 'C:/'
fileName = os.listdir(file_dir)
new_workbook = xlwt.Workbook()
workSheet = new_workbook.add_sheet('new_test')
n = 0
for i in fileName:
    workSheet.write(n, 0, i)
    n += 1
new_workbook.save('D:/file_20200609.xls')
print(type(fileName))

# 问题
# 1、如何判断文件还是文件夹？


# ------2、excel渲染一张图片 -------#
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color
from openpyxl.styles import PatternFill
from PIL import Image

workbook = Workbook()
worksheet = workbook.active
im = Image.open("asd.png")
im_width = im.size[0]
im_height = im.size[1]
pix = im.load()
for row in range(1, im_height):
    for col in range(1, im_width):
        cell = worksheet.cell(column=col, row=row)
        pixpoint = pix[col - 1, row - 1]
        pixColor = "FF%02X%02X%02X" % (pixpoint[0], pixpoint[1], pixpoint[2])
        fill = PatternFill(patternType='solid', fgColor=Color(rgb=pixColor))
        cell.fill = fill
    worksheet.row_dimensions[row].height = 6
for col in range(1, im_width):
    worksheet.column_dimensions[get_column_letter(col)].width = 1
workbook.save('D:/imageFile.xls')

#注意：
# 1、安装包：openpyxl，PIL（python3.8用pillow代替PIL）
#  pip install -index-url=https://pypi.tuna.tsinghua.edu.cn/simple  Pillow
# 2、图片不能太大，报错int