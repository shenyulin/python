import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl.utils import get_column_letter

#------一、读excel数据--------#
#1、读取工作簿 open_workbook
#2、读取工作表 sheet_by_index
#3、读取某一行列中的数据
excelFile =xlrd.open_workbook("G:\\test.xlsx")
excelSheet=excelFile.sheet_by_index(0)
msg=excelSheet.row(0)[0].value
#print(msg)


#--------二、写excel数据-------#
#1、新建工作薄 Workbook
#2、新建一个工作表 add_sheet
#3、在表中写入数据 write
#4、保存表 save
writeFile=xlwt.Workbook()
writeSheet=writeFile.add_sheet("test")
writeSheet.write(0,0,"name")
writeSheet.write(0,1,"sex")
writeFile.save("G:\\newFile.xls")


#--------三、拷贝文件--------#
#1、读取工作薄、工作表
readFile=xlrd.open_workbook("G:\\newFile.xls",formatting_info=True)
readSheet=readFile.sheet_by_index(0)

newFile=copy(readFile)
newFileSheet=newFile.get_sheet(0)
newFileSheet.write(1,1,"你还好么？")
newFileSheet.write(1,2,"还好")
newFileSheet.write(2,1,"你吃了么")
newFileSheet.write(2,2,"吃了")
newFile.save("G:\\newFile1.xls")


#-------------------------作业-----------------------#
#1、填入任意数字，宋体、左、上对齐
dayFile=xlrd.open_workbook("G:\\日统计.xls",formatting_info=True)
daySheet=dayFile.sheet_by_index(0)

#2、拷贝工作薄、工作表
newDayFile=copy(dayFile)
newDaySheet=newDayFile.get_sheet(0)

#3、设置格式
style=xlwt.XFStyle()

#设置字体和对齐
font=xlwt.Font()
font.name='宋体'
style.font=font

alignment=xlwt.Alignment()
alignment.horz=xlwt.Alignment.HORZ_LEFT
alignment.vert=xlwt.Alignment.VERT_TOP
style.alignment=alignment
newDaySheet.write(1,1,222,style)
newDayFile.save("G:\\日统计_new.xls")


#修改文件字体
font=xlwt.Font()
font.name='隶书'
font.colour_index=2
style.font=font

#获取表行数、列数
rowsNum=daySheet.nrows
colsNum=daySheet.ncols

#遍历表格数据
for i in range(rowsNum):
    for  j in range(colsNum): 
        txt=daySheet.cell(i,j).value
        if ((type(txt).__name__ ==  'float') and txt > 10):
            newDaySheet.write(i,j,txt,style)
        else:
            pass

newDayFile.save("G:\\日统计_new.xls")











